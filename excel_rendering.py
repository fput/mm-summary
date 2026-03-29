"""Excel workbook rendering."""

# Sheet layout:
#
#   A          B     C    …  M     N       O
#   [Year]     Jan   Feb  …  Dec   Total   Average
#   Income                                 (avg label)
#     Salary   100        …
#   Total Inc  ───        …
#   Expenses
#     Group A
#       Cat1   50         …
#   Total Exp  ───        …
#   Savings    =Inc-Exp   …
#   Savings %  =Sav/Inc   …

from __future__ import annotations

import datetime as dt
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Literal

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series  # pyright: ignore
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from aggregation import (
    average_context,
    build_summary_data,
    category_sort_key,
    ordered_expense_groups,
)
from models import Money, Month, SummaryData, TaxonomyConfig, Transaction
from user_config import DEFAULT_TAXONOMY

__all__ = ["create_workbook"]

LABEL_COL = 1
MONTH_START_COL = 2
MONTH_END_COL = 13
TOTAL_COL = 14
AVG_COL = 15
MAX_COL = 15

CURRENCY_NUMBER_FORMAT = "#,##0;-#,##0;;@"
PERCENT_NUMBER_FORMAT = "0.0%"

MONTHS = tuple(range(1, 13))
MONTH_ABBR_EN = (
    "Jan",
    "Feb",
    "Mar",
    "Apr",
    "May",
    "Jun",
    "Jul",
    "Aug",
    "Sep",
    "Oct",
    "Nov",
    "Dec",
)

RAW_TRANSACTION_COLUMNS = (
    "bookingDate",
    "valueDate",
    "amount",
    "currency",
    "category",
    "name",
    "purpose",
    "comment",
    "bookingText",
    "accountNumber",
    "bankCode",
    "transactionCode",
    "purposeCode",
    "bookingKey",
    "primanotaNumber",
    "batchReference",
    "endToEndReference",
    "mandateReference",
    "creditorId",
    "returnReason",
    "booked",
    "checkmark",
    "id",
)

COLOR_WHITE = "FFFFFF"
COLOR_INCOME_HEADER = "A9D18E"
COLOR_INCOME_ROW = "E2F0D9"
COLOR_INCOME_ROW_RIGHT = "C6DBB8"
COLOR_INCOME_TOTAL = "C6E0B4"
COLOR_INCOME_TOTAL_RIGHT = "A0C98A"

COLOR_EXPENSE_HEADER = "F4B084"
COLOR_EXPENSE_GROUP = "F8CBAD"
COLOR_EXPENSE_ROW = "FCE4D6"
COLOR_EXPENSE_ROW_RIGHT = "F2CDB7"
COLOR_EXPENSE_TOTAL = "F4B183"
COLOR_EXPENSE_TOTAL_RIGHT = "F4B183"
COLOR_EXPENSE_HEATMAP_LOW = COLOR_EXPENSE_ROW
COLOR_EXPENSE_HEATMAP_HIGH = COLOR_EXPENSE_TOTAL

COLOR_POSITIVE = "C6E0B4"
COLOR_NEGATIVE = "F4CCCC"

FONT_SIZE_NORMAL = 15
FONT_SIZE_HEADER = 18
ROW_HEIGHT_SECTION = 20
ROW_HEIGHT_BOTTOM = 26

LIGHT_BORDER_SIDE = Side(style="thin", color="D0D0D0")
STRONG_BORDER_SIDE = Side(style="thin", color="606060")


@dataclass(frozen=True)
class SheetStyles:
    font_normal: Font
    font_bold: Font
    font_header: Font
    fill_white: PatternFill
    fill_income_header: PatternFill
    fill_income_row: PatternFill
    fill_income_row_right: PatternFill
    fill_income_total: PatternFill
    fill_income_total_right: PatternFill
    fill_expense_header: PatternFill
    fill_expense_group: PatternFill
    fill_expense_row: PatternFill
    fill_expense_row_right: PatternFill
    fill_expense_total: PatternFill
    fill_expense_total_right: PatternFill
    fill_positive: PatternFill
    fill_negative: PatternFill


def create_workbook(
    raw_transactions: list[dict],
    transactions: list[Transaction],
    *,
    year: int,
    output_path: Path,
    add_comments: bool,
    include_transactions_sheet: bool,
    expense_heatmap: bool = True,
    taxonomy: TaxonomyConfig = DEFAULT_TAXONOMY,
) -> None:
    wb = Workbook()
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    styles = _build_sheet_styles()
    summary_ws = wb.worksheets[0]
    summary_ws.title = "Summary"
    SummarySheetRenderer(
        summary_ws,
        model=build_summary_data(transactions, taxonomy=taxonomy),
        year=year,
        add_comments=add_comments,
        styles=styles,
        expense_heatmap=expense_heatmap,
        taxonomy=taxonomy,
    ).render()

    if include_transactions_sheet:
        _make_transactions_sheet(
            wb.create_sheet("Transactions"),
            raw_transactions,
            styles=styles,
        )

    wb.save(output_path)


def _displayed_amount(amount: Money, kind: Literal["income", "expense"]) -> Money:
    # Expenses are stored as positive values internally but displayed as
    # positive costs in the sheet (not negative income).
    return amount if kind == "income" else -amount


def _build_sheet_styles() -> SheetStyles:
    return SheetStyles(
        font_normal=Font(size=FONT_SIZE_NORMAL),
        font_bold=Font(size=FONT_SIZE_NORMAL, bold=True),
        font_header=Font(size=FONT_SIZE_HEADER, bold=True),
        fill_white=PatternFill(fill_type="solid", fgColor=COLOR_WHITE),
        fill_income_header=PatternFill(fill_type="solid", fgColor=COLOR_INCOME_HEADER),
        fill_income_row=PatternFill(fill_type="solid", fgColor=COLOR_INCOME_ROW),
        fill_income_row_right=PatternFill(
            fill_type="solid", fgColor=COLOR_INCOME_ROW_RIGHT
        ),
        fill_income_total=PatternFill(fill_type="solid", fgColor=COLOR_INCOME_TOTAL),
        fill_income_total_right=PatternFill(
            fill_type="solid", fgColor=COLOR_INCOME_TOTAL_RIGHT
        ),
        fill_expense_header=PatternFill(
            fill_type="solid", fgColor=COLOR_EXPENSE_HEADER
        ),
        fill_expense_group=PatternFill(fill_type="solid", fgColor=COLOR_EXPENSE_GROUP),
        fill_expense_row=PatternFill(fill_type="solid", fgColor=COLOR_EXPENSE_ROW),
        fill_expense_row_right=PatternFill(
            fill_type="solid", fgColor=COLOR_EXPENSE_ROW_RIGHT
        ),
        fill_expense_total=PatternFill(fill_type="solid", fgColor=COLOR_EXPENSE_TOTAL),
        fill_expense_total_right=PatternFill(
            fill_type="solid", fgColor=COLOR_EXPENSE_TOTAL_RIGHT
        ),
        fill_positive=PatternFill(
            fill_type="solid",
            start_color=COLOR_POSITIVE,
            end_color=COLOR_POSITIVE,
        ),
        fill_negative=PatternFill(
            fill_type="solid",
            start_color=COLOR_NEGATIVE,
            end_color=COLOR_NEGATIVE,
        ),
    )


def _set_border(cell, *, left=None, right=None, top=None, bottom=None) -> None:
    # openpyxl replaces the entire Border object, so we have to
    # preserve sides we're not changing
    cell.border = Border(
        left=cell.border.left if left is None else left,
        right=cell.border.right if right is None else right,
        top=cell.border.top if top is None else top,
        bottom=cell.border.bottom if bottom is None else bottom,
    )


def _set_row_vertical_center(ws, row: int, max_col: int = MAX_COL) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.alignment = Alignment(
            horizontal=cell.alignment.horizontal,
            vertical="center",
            text_rotation=cell.alignment.textRotation,
            wrap_text=cell.alignment.wrapText,
            shrink_to_fit=cell.alignment.shrinkToFit,
            indent=cell.alignment.indent,
        )


def _apply_row_fills(
    ws,
    row: int,
    *,
    base_fill: PatternFill,
    right_fill: PatternFill,
    max_col: int = MAX_COL,
) -> None:
    for col in range(1, max_col + 1):
        ws.cell(row=row, column=col).fill = (
            right_fill if col in (TOTAL_COL, AVG_COL) else base_fill
        )


def _set_cell_value_or_blank(cell, value) -> None:
    cell.value = None if value in (None, 0) else value


def _apply_table_borders(
    ws,
    *,
    header_row: int,
    last_row: int,
    expense_total_row: int,
    savings_row: int,
) -> None:
    for row in range(header_row, last_row + 1):
        for col in range(1, MAX_COL + 1):
            cell = ws.cell(row=row, column=col)
            _set_border(
                cell,
                left=LIGHT_BORDER_SIDE,
                right=LIGHT_BORDER_SIDE,
                top=LIGHT_BORDER_SIDE,
                bottom=LIGHT_BORDER_SIDE,
            )

    for row in range(header_row, last_row + 1):
        _set_border(ws.cell(row=row, column=LABEL_COL), left=STRONG_BORDER_SIDE)
        _set_border(ws.cell(row=row, column=MAX_COL), right=STRONG_BORDER_SIDE)

    for col in range(1, MAX_COL + 1):
        _set_border(ws.cell(row=header_row, column=col), top=STRONG_BORDER_SIDE)
        _set_border(ws.cell(row=last_row, column=col), bottom=STRONG_BORDER_SIDE)
        _set_border(
            ws.cell(row=expense_total_row, column=col), bottom=STRONG_BORDER_SIDE
        )
        _set_border(ws.cell(row=savings_row, column=col), top=STRONG_BORDER_SIDE)


def _add_sign_fill_rules(ws, cell_range: str, styles: SheetStyles) -> None:
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            fill=styles.fill_positive,
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="lessThan",
            formula=["0"],
            fill=styles.fill_negative,
        ),
    )


def _add_summary_chart_from_table(
    ws, *, header_row: int, last_row: int, year: int, anchor_cell: str
) -> None:
    point_count = max(1, last_row - header_row)

    chart = BarChart()
    chart.type = "bar"
    chart.style = 1
    chart.title = f"Categories / Average {year}"
    chart.x_axis.title = "EUR"
    chart.y_axis.title = "Category"
    chart.legend = None
    chart.varyColors = False
    chart.height = max(18, min(40, 6 + point_count * 0.6))
    chart.width = 25
    chart.x_axis.scaling.orientation = "maxMin"
    chart.x_axis.tickLblPos = "nextTo"
    chart.x_axis.tickLblSkip = 1

    chart.add_data(
        Reference(ws, min_col=AVG_COL, min_row=header_row, max_row=last_row),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(ws, min_col=LABEL_COL, min_row=header_row + 1, max_row=last_row)
    )
    ws.add_chart(chart, anchor_cell)


def _add_monthly_income_expense_chart(
    ws,
    *,
    header_row: int,
    income_total_row: int,
    expense_total_row: int,
    year: int,
    anchor_cell: str,
) -> None:
    chart = BarChart()
    chart.type = "col"
    chart.style = 1
    chart.grouping = "clustered"
    chart.overlap = 0
    chart.title = f"Monthly Income and Expenses {year}"
    chart.x_axis.title = "Month"
    chart.y_axis.title = "EUR"
    chart.legend = None
    chart.height = 10
    chart.width = 25

    income_series = Series(
        Reference(
            ws,
            min_col=MONTH_START_COL,
            max_col=MONTH_END_COL,
            min_row=income_total_row,
            max_row=income_total_row,
        ),
        title="Income",
    )
    expense_series = Series(
        Reference(
            ws,
            min_col=MONTH_START_COL,
            max_col=MONTH_END_COL,
            min_row=expense_total_row,
            max_row=expense_total_row,
        ),
        title="Expenses",
    )

    income_series.graphicalProperties.solidFill = "70AD47"
    income_series.graphicalProperties.line.solidFill = "70AD47"
    expense_series.graphicalProperties.solidFill = "C00000"
    expense_series.graphicalProperties.line.solidFill = "C00000"

    chart.append(income_series)
    chart.append(expense_series)
    chart.set_categories(
        Reference(
            ws,
            min_col=MONTH_START_COL,
            max_col=MONTH_END_COL,
            min_row=header_row,
            max_row=header_row,
        )
    )
    ws.add_chart(chart, anchor_cell)


class SummarySheetRenderer:
    """Render the summary sheet from aggregated yearly data."""

    def __init__(
        self,
        ws,
        *,
        model: SummaryData,
        year: int,
        add_comments: bool,
        styles: SheetStyles,
        expense_heatmap: bool,
        taxonomy: TaxonomyConfig,
    ) -> None:
        self.ws = ws
        self.model = model
        self.year = year
        self.add_comments = add_comments
        self.styles = styles
        self.expense_heatmap = expense_heatmap
        self.taxonomy = taxonomy
        self.header_row = 1
        self.income_total_row = 0
        self.expense_total_row = 0
        self.expense_body_start_row = 0
        self.savings_row = 0
        self.savings_rate_row = 0
        self.row = self.header_row
        # For the current (incomplete) year, average over elapsed months only.
        self.avg_month_count, self.avg_label = average_context(year)

    def render(self) -> None:
        self._render_header_row()
        self.income_total_row = self._render_income_block()
        self.expense_total_row = self._render_expense_block()
        if self.expense_heatmap:
            self._add_expense_heatmap()
        self._render_savings_section()
        self._finalize_layout()
        self._add_charts()

    def _cell(self, row: int, col: int):
        return self.ws.cell(row=row, column=col)

    def _set_font(self, cell, *, bold: bool = False, header: bool = False) -> None:
        if header:
            cell.font = self.styles.font_header
        elif bold:
            cell.font = self.styles.font_bold
        else:
            cell.font = self.styles.font_normal

    def _cell_ref(self, row: int, col: int) -> str:
        return f"{get_column_letter(col)}{row}"

    def _write_money_formula(self, cell, formula: str, bold: bool = False) -> None:
        cell.value = formula
        cell.number_format = CURRENCY_NUMBER_FORMAT
        self._set_font(cell, bold=bold)

    def _write_row_total_and_average(
        self, row: int, *, whole_row_bold: bool = False, total_avg_bold: bool = False
    ) -> None:
        total_cell = self._cell(row, TOTAL_COL)
        avg_cell = self._cell(row, AVG_COL)
        total_ref = self._cell_ref(row, TOTAL_COL)

        self._write_money_formula(
            total_cell,
            f"=SUM({self._cell_ref(row, MONTH_START_COL)}:{self._cell_ref(row, MONTH_END_COL)})",
            bold=(whole_row_bold or total_avg_bold),
        )
        self._write_money_formula(
            avg_cell,
            f"={total_ref}/{self.avg_month_count}",
            bold=(whole_row_bold or total_avg_bold),
        )

    def _render_header_row(self) -> None:
        for col, value in enumerate(
            [self.year, *MONTH_ABBR_EN, "Total", "Average"], start=1
        ):
            cell = self._cell(self.header_row, col)
            cell.value = value
            self._set_font(cell, header=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = self.styles.fill_white
            if col == LABEL_COL:
                cell.number_format = "0"

        self.row = self.header_row + 1

    def _write_label_row(
        self,
        label: str,
        *,
        base_fill: PatternFill,
        bold: bool = True,
        indent: int = 0,
        extra_text_col: int | None = None,
        extra_text: str | None = None,
    ) -> int:
        row = self.row
        label_cell = self._cell(row, LABEL_COL)
        label_cell.value = label
        label_cell.alignment = Alignment(indent=indent)
        self._set_font(label_cell, bold=bold)

        if extra_text_col is not None and extra_text:
            extra_cell = self._cell(row, extra_text_col)
            extra_cell.value = extra_text
            extra_cell.alignment = Alignment(horizontal="center", vertical="center")
            self._set_font(extra_cell, bold=True)

        _apply_row_fills(self.ws, row, base_fill=base_fill, right_fill=base_fill)
        self.ws.row_dimensions[row].height = ROW_HEIGHT_SECTION
        _set_row_vertical_center(self.ws, row)
        self.row += 1
        return row

    def _write_data_row(
        self,
        label: str,
        *,
        base_fill: PatternFill,
        right_fill: PatternFill,
        month_values: dict[Month, Money],
        kind: Literal["income", "expense"],
        indent: int = 0,
        label_bold: bool = False,
        whole_row_bold: bool = False,
        total_avg_bold: bool = False,
        note_group: str | None = None,
        note_category: str | None = None,
    ) -> int:
        row = self.row
        label_cell = self._cell(row, LABEL_COL)
        label_cell.value = label
        label_cell.alignment = Alignment(indent=indent)
        self._set_font(label_cell, bold=(label_bold or whole_row_bold))

        for month in MONTHS:
            col = MONTH_START_COL + month - 1
            cell = self._cell(row, col)
            shown = _displayed_amount(month_values.get(month, Decimal("0")), kind)
            _set_cell_value_or_blank(cell, shown)
            cell.number_format = CURRENCY_NUMBER_FORMAT
            self._set_font(cell, bold=whole_row_bold)
            if self.add_comments and note_group and note_category and shown != 0:
                note = self.model.notes.get((note_group, note_category, month))
                if note:
                    cell.comment = Comment(note, "MoneyMoney")

        self._write_row_total_and_average(
            row,
            whole_row_bold=whole_row_bold,
            total_avg_bold=total_avg_bold,
        )
        _apply_row_fills(self.ws, row, base_fill=base_fill, right_fill=right_fill)
        self.row += 1
        return row

    def _write_sum_row(
        self,
        label: str,
        *,
        start_row: int,
        end_row: int,
        base_fill: PatternFill,
        right_fill: PatternFill,
        label_bold: bool = False,
        whole_row_bold: bool = False,
        total_avg_bold: bool = False,
    ) -> int:
        row = self.row
        label_cell = self._cell(row, LABEL_COL)
        label_cell.value = label
        self._set_font(label_cell, bold=(label_bold or whole_row_bold))

        for month in MONTHS:
            col = MONTH_START_COL + month - 1
            cell = self._cell(row, col)
            if start_row > end_row:
                cell.value = None
                cell.number_format = CURRENCY_NUMBER_FORMAT
                self._set_font(cell, bold=whole_row_bold)
                continue
            self._write_money_formula(
                cell,
                f"=SUM({self._cell_ref(start_row, col)}:{self._cell_ref(end_row, col)})",
                bold=whole_row_bold,
            )

        self._write_row_total_and_average(
            row,
            whole_row_bold=whole_row_bold,
            total_avg_bold=total_avg_bold,
        )
        _apply_row_fills(self.ws, row, base_fill=base_fill, right_fill=right_fill)
        self.row += 1
        return row

    def _render_income_block(self) -> int:
        self._write_label_row(
            "Income",
            base_fill=self.styles.fill_income_header,
            extra_text_col=AVG_COL,
            extra_text=self.avg_label,
        )
        start_row = self.row

        for category in sorted(
            self.model.income_categories,
            key=lambda name: category_sort_key(
                self.taxonomy.income_group,
                name,
                category_priority_by_group=self.taxonomy.category_priority_by_group,
            ),
        ):
            self._write_data_row(
                category,
                base_fill=self.styles.fill_income_row,
                right_fill=self.styles.fill_income_row_right,
                month_values=self.model.income_categories[category],
                kind="income",
                indent=1,
                note_group=self.taxonomy.income_group,
                note_category=category,
            )

        return self._write_sum_row(
            "Total Income",
            start_row=start_row,
            end_row=self.row - 1,
            base_fill=self.styles.fill_income_total,
            right_fill=self.styles.fill_income_total_right,
            label_bold=True,
            whole_row_bold=True,
            total_avg_bold=True,
        )

    def _render_expense_block(self) -> int:
        self._write_label_row(
            "Expenses",
            base_fill=self.styles.fill_expense_header,
        )
        start_row = self.row
        self.expense_body_start_row = start_row

        for group in ordered_expense_groups(
            self.model.expense_groups,
            expense_group_order=self.taxonomy.expense_group_order,
        ):
            self._write_label_row(group, base_fill=self.styles.fill_expense_group)
            categories = self.model.expense_groups[group]
            for category in sorted(
                categories,
                key=lambda name: category_sort_key(
                    group,
                    name,
                    category_priority_by_group=self.taxonomy.category_priority_by_group,
                ),
            ):
                self._write_data_row(
                    category,
                    base_fill=self.styles.fill_expense_row,
                    right_fill=self.styles.fill_expense_row_right,
                    month_values=categories[category],
                    kind="expense",
                    indent=1,
                    note_group=group,
                    note_category=category,
                )

        return self._write_sum_row(
            "Total Expenses",
            start_row=start_row,
            end_row=self.row - 1,
            base_fill=self.styles.fill_expense_total,
            right_fill=self.styles.fill_expense_total_right,
            label_bold=True,
            whole_row_bold=True,
            total_avg_bold=True,
        )

    def _add_expense_heatmap(self) -> None:
        if self.expense_body_start_row >= self.expense_total_row:
            return

        expense_heatmap_range = (
            f"{self._cell_ref(self.expense_body_start_row, MONTH_START_COL)}:"
            f"{self._cell_ref(self.expense_total_row - 1, MONTH_END_COL)}"
        )
        self.ws.conditional_formatting.add(
            expense_heatmap_range,
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color=COLOR_EXPENSE_HEATMAP_LOW,
                end_type="max",
                end_color=COLOR_EXPENSE_HEATMAP_HIGH,
            ),
        )

    def _render_savings_section(self) -> None:
        self._write_savings_row()
        self._write_savings_rate_row()

        _apply_table_borders(
            self.ws,
            header_row=self.header_row,
            last_row=self.savings_rate_row,
            expense_total_row=self.expense_total_row,
            savings_row=self.savings_row,
        )
        self.row += 1

    def _write_savings_row(self) -> None:
        self.savings_row = self.row
        label_cell = self._cell(self.row, LABEL_COL)
        label_cell.value = "Savings"
        label_cell.fill = self.styles.fill_white
        self._set_font(label_cell, bold=False)

        for month in MONTHS:
            col = MONTH_START_COL + month - 1
            cell = self._cell(self.row, col)
            self._write_money_formula(
                cell,
                f"={self._cell_ref(self.income_total_row, col)}-{self._cell_ref(self.expense_total_row, col)}",
            )
            cell.fill = self.styles.fill_white

        total_cell = self._cell(self.row, TOTAL_COL)
        avg_cell = self._cell(self.row, AVG_COL)
        self._write_money_formula(
            total_cell,
            f"={self._cell_ref(self.income_total_row, TOTAL_COL)}-{self._cell_ref(self.expense_total_row, TOTAL_COL)}",
            bold=True,
        )
        self._write_money_formula(
            avg_cell,
            f"={self._cell_ref(self.income_total_row, AVG_COL)}-{self._cell_ref(self.expense_total_row, AVG_COL)}",
            bold=True,
        )
        total_cell.fill = self.styles.fill_white
        avg_cell.fill = self.styles.fill_white

        savings_range = (
            f"{self._cell_ref(self.row, MONTH_START_COL)}:"
            f"{self._cell_ref(self.row, AVG_COL)}"
        )
        _add_sign_fill_rules(self.ws, savings_range, styles=self.styles)

        self.ws.row_dimensions[self.row].height = ROW_HEIGHT_BOTTOM
        _set_row_vertical_center(self.ws, self.row)
        self.row += 1

    def _write_savings_rate_row(self) -> None:
        self.savings_rate_row = self.row
        label_cell = self._cell(self.row, LABEL_COL)
        label_cell.value = "Savings Rate"
        label_cell.fill = self.styles.fill_white
        self._set_font(label_cell, bold=True)

        for month in MONTHS:
            col = MONTH_START_COL + month - 1
            income_ref = self._cell_ref(self.income_total_row, col)
            savings_ref = self._cell_ref(self.savings_row, col)
            cell = self._cell(self.row, col)
            cell.value = f'=IF({income_ref}=0,"",{savings_ref}/{income_ref})'
            cell.number_format = PERCENT_NUMBER_FORMAT
            cell.fill = self.styles.fill_white
            cell.alignment = Alignment(horizontal="center")
            self._set_font(cell, bold=True)

        merged_cell = self._cell(self.row, TOTAL_COL)
        merged_cell.value = (
            f'=IF({self._cell_ref(self.income_total_row, TOTAL_COL)}=0,"",'
            f"{self._cell_ref(self.savings_row, TOTAL_COL)}/{self._cell_ref(self.income_total_row, TOTAL_COL)})"
        )
        merged_cell.number_format = PERCENT_NUMBER_FORMAT
        merged_cell.fill = self.styles.fill_white
        merged_cell.alignment = Alignment(horizontal="center", vertical="center")
        self._set_font(merged_cell, bold=True)
        self._cell(self.row, AVG_COL).fill = self.styles.fill_white

        savings_rate_range = (
            f"{self._cell_ref(self.row, MONTH_START_COL)}:"
            f"{self._cell_ref(self.row, TOTAL_COL)}"
        )
        _add_sign_fill_rules(self.ws, savings_rate_range, styles=self.styles)

        self.ws.row_dimensions[self.row].height = ROW_HEIGHT_BOTTOM
        _set_row_vertical_center(self.ws, self.row)
        self.ws.merge_cells(
            start_row=self.row,
            start_column=TOTAL_COL,
            end_row=self.row,
            end_column=AVG_COL,
        )

    def _finalize_layout(self) -> None:
        for col in range(1, MAX_COL + 1):
            width = 44 if col == LABEL_COL else 18 if col == AVG_COL else 15
            self.ws.column_dimensions[get_column_letter(col)].width = width

        self.ws.freeze_panes = "B2"
        self.ws.sheet_view.showGridLines = True

    def _add_charts(self) -> None:
        chart_anchor_row = self.savings_rate_row + 3
        _add_summary_chart_from_table(
            self.ws,
            header_row=self.header_row,
            last_row=self.expense_total_row,
            year=self.year,
            anchor_cell=f"A{chart_anchor_row}",
        )
        _add_monthly_income_expense_chart(
            self.ws,
            header_row=self.header_row,
            income_total_row=self.income_total_row,
            expense_total_row=self.expense_total_row,
            year=self.year,
            anchor_cell=f"H{chart_anchor_row}",
        )


def _make_transactions_sheet(
    ws, raw_transactions: list[dict], styles: SheetStyles
) -> None:
    ws.append(list(RAW_TRANSACTION_COLUMNS))

    for raw in raw_transactions:
        row = []
        for column in RAW_TRANSACTION_COLUMNS:
            value = raw.get(column)
            if isinstance(value, dt.datetime):
                value = value.replace(tzinfo=None)
            row.append(value)
        ws.append(row)

    ws.freeze_panes = "A2"

    for index, column_name in enumerate(RAW_TRANSACTION_COLUMNS, start=1):
        width = (
            24
            if column_name in {"purpose", "comment", "name", "category", "bookingText"}
            else max(len(column_name) + 2, 14)
        )
        ws.column_dimensions[get_column_letter(index)].width = width

        cell = ws.cell(row=1, column=index)
        cell.font = styles.font_header
        cell.fill = styles.fill_white
        cell.alignment = Alignment(horizontal="center", vertical="center")
